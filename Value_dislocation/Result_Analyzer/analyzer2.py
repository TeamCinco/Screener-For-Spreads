"""
Opportunity Analyzer - Rotation Aware Version
Ranks stocks using:
• Technical dislocation
• Valuation
• Risk metrics
• ETF capital flow alignment
"""

import pandas as pd
import numpy as np
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ============================================================================
# FILE PATHS
# ============================================================================

INPUT_FILE = "/Users/jazzhashzzz/Desktop/Screener For Spreads/output/screener_results_enhanced.xlsx"
ETF_CATEGORY_FILE = "/Users/jazzhashzzz/Desktop/Screener For Spreads/output/2026-02-18/category_composite.csv"
OUTPUT_FILE = "/Users/jazzhashzzz/Desktop/Screener For Spreads/output/opportunities_ranked2.xlsx"

# ============================================================================
# SECTOR → ETF CATEGORY MAP
# ============================================================================

SECTOR_TO_CATEGORY = {
    "Technology": ["Technology", "Large Growth", "Large Blend"],
    "Financial Services": ["Financial", "Large Value"],
    "Energy": ["Commodities Focused", "Energy Limited Partnership"],
    "Basic Materials": ["Equity Precious Metals", "Commodities Broad Basket"],
    "Real Estate": ["Large Value", "Large Blend"],
    "Healthcare": ["Large Growth", "Large Blend"],
    "Industrials": ["Large Blend", "Mid-Cap Blend"],
    "Consumer Cyclical": ["Mid-Cap Blend", "Large Blend"],
    "Consumer Defensive": ["Large Value", "Large Blend"],
    "Communication Services": ["Large Growth", "Large Blend"],
}

# ============================================================================
# SCORING CONFIG
# ============================================================================

CRITERIA = {
    'z_score': {'optimal': (-3.0, -2.0), 'acceptable': (-4.0, -1.5), 'weight': 0.20},
    'pe_ratio': {'optimal': (5, 25), 'acceptable': (0, 40), 'weight': 0.15},
    'drop_from_high_pct': {'optimal': (-40, -20), 'acceptable': (-60, -10), 'weight': 0.10},
    'p10': {'optimal': (-40, -15), 'acceptable': (-60, -10), 'weight': 0.15},
    'volatility': {'optimal': (25, 50), 'acceptable': (15, 70), 'weight': 0.10},
    'vol_skew_ratio': {'optimal': (0.8, 1.1), 'acceptable': (0.6, 1.5), 'weight': 0.10},
    'risk_state_score': {'optimal': (20, 50), 'acceptable': (10, 70), 'weight': 0.10},
    'volume_surge': {'optimal': (1.3, 3.0), 'acceptable': (0.8, 5.0), 'weight': 0.10},
}

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def score_metric(value, optimal, acceptable):
    if pd.isna(value):
        return 0
    opt_min, opt_max = optimal
    acc_min, acc_max = acceptable

    if opt_min <= value <= opt_max:
        return 100

    if acc_min <= value <= acc_max:
        if value < opt_min:
            distance = opt_min - value
            max_dist = opt_min - acc_min
        else:
            distance = value - opt_max
            max_dist = acc_max - opt_max
        return max(0, 100 - (distance / max_dist * 50))

    return 0


def calculate_composite_score(row):
    total = 0
    weight_total = 0
    for metric, params in CRITERIA.items():
        if metric in row and pd.notna(row[metric]):
            s = score_metric(row[metric], params['optimal'], params['acceptable'])
            total += s * params['weight']
            weight_total += params['weight']
    return total / weight_total if weight_total > 0 else 0


# ============================================================================
# CAPITAL FLOW INTEGRATION
# ============================================================================

def normalize_series(series):
    """Normalize ETF composite to -1 to 1 scale"""
    max_abs = series.abs().max()
    if max_abs == 0:
        return series
    return series / max_abs


def build_category_strength():
    etf_df = pd.read_csv(ETF_CATEGORY_FILE)
    etf_df["normalized"] = normalize_series(etf_df["avg_composite"])
    return dict(zip(etf_df["category"], etf_df["normalized"]))


def get_sector_capital_score(sector, category_strength):
    if sector not in SECTOR_TO_CATEGORY:
        return 0

    categories = SECTOR_TO_CATEGORY[sector]
    scores = [category_strength.get(cat, 0) for cat in categories]

    return np.mean(scores) if len(scores) > 0 else 0


# ============================================================================
# MAIN ANALYSIS
# ============================================================================

def analyze_opportunities():

    print("\nRotation Aware Opportunity Analyzer\n")

    df = pd.read_excel(INPUT_FILE, sheet_name='All Results')
    print(f"Loaded {len(df)} stocks")

    # Quality filters
    df = df[df['pe_ratio'].notna() & (df['pe_ratio'] > 0)]
    df = df[df['z_score'].notna() & (df['z_score'] < -1.5)]
    df = df[df['drop_from_high_pct'] > -70]

    print(f"{len(df)} passed filters")

    if len(df) == 0:
        print("No valid candidates.")
        return

    # Technical score
    df['opportunity_score'] = df.apply(calculate_composite_score, axis=1)

    # Capital score
    category_strength = build_category_strength()
    df['capital_flow_score'] = df['sector'].apply(
        lambda x: get_sector_capital_score(x, category_strength)
    )

    # Final score blend
    df['final_score'] = (
        df['opportunity_score'] * 0.75 +
        df['capital_flow_score'] * 25
    )

    df = df.sort_values('final_score', ascending=False)

    # Tier
    df['tier'] = df['final_score'].apply(
        lambda x: 'STRONG' if x >= 70 else ('REVIEW' if x >= 50 else 'PASS')
    )

    # ============================================================================
    # SAVE OUTPUT
    # ============================================================================

    df.to_excel(OUTPUT_FILE, index=False)

    # Apply simple formatting
    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    for col in range(1, ws.max_column + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.column_dimensions[get_column_letter(col)].width = 18

    for row in range(2, ws.max_row + 1):
        tier_cell = ws.cell(row=row, column=list(df.columns).index("tier")+1)
        capital_cell = ws.cell(row=row, column=list(df.columns).index("capital_flow_score")+1)

        if tier_cell.value == "STRONG":
            tier_cell.fill = PatternFill("solid", fgColor="C6EFCE")
        elif tier_cell.value == "REVIEW":
            tier_cell.fill = PatternFill("solid", fgColor="FFEB9C")
        else:
            tier_cell.fill = PatternFill("solid", fgColor="FFC7CE")

        if capital_cell.value > 0.3:
            capital_cell.fill = PatternFill("solid", fgColor="C6EFCE")
        elif capital_cell.value < -0.3:
            capital_cell.fill = PatternFill("solid", fgColor="FFC7CE")

    wb.save(OUTPUT_FILE)

    print("\nSaved:", OUTPUT_FILE)
    print("\nTop 10:")
    print(df[['ticker','sector','opportunity_score',
              'capital_flow_score','final_score','tier']].head(10).to_string(index=False))


# ============================================================================
# RUN
# ============================================================================

if __name__ == "__main__":
    analyze_opportunities()
