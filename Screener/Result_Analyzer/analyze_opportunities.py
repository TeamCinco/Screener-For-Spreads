"""
Opportunity Analyzer - Scores and ranks screener output
Color-coded Excel output with new risk columns
"""
import pandas as pd
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================================
# CONFIGURATION
# ============================================================================

INPUT_FILE = "/Users/jazzhashzzz/Desktop/Screener For Spreads/output/screener_results.xlsx"
OUTPUT_FILE = "/Users/jazzhashzzz/Desktop/Screener For Spreads/output/opportunities_ranked.xlsx"

CRITERIA = {
    'z_score': {
        'optimal_range': (-3.0, -2.0),
        'acceptable_range': (-4.0, -1.5),
        'weight': 0.20
    },
    'pe_ratio': {
        'optimal_range': (5, 25),
        'acceptable_range': (0, 40),
        'weight': 0.15
    },
    'drop_from_high_pct': {
        'optimal_range': (-40, -20),
        'acceptable_range': (-60, -10),
        'weight': 0.10
    },
    'p10': {
        'optimal_range': (-40, -15),
        'acceptable_range': (-60, -10),
        'weight': 0.15
    },
    'volatility': {
        'optimal_range': (25, 50),
        'acceptable_range': (15, 70),
        'weight': 0.10
    },
    'vol_skew_ratio': {
        'optimal_range': (0.8, 1.1),
        'acceptable_range': (0.6, 1.5),
        'weight': 0.10
    },
    'risk_state_score': {
        'optimal_range': (20, 50),
        'acceptable_range': (10, 70),
        'weight': 0.10
    },
    'volume_surge': {
        'optimal_range': (1.3, 3.0),
        'acceptable_range': (0.8, 5.0),
        'weight': 0.10
    },
}

# ============================================================================
# SCORING
# ============================================================================

def score_metric(value, optimal_range, acceptable_range):
    if pd.isna(value):
        return 0
    opt_min, opt_max = optimal_range
    acc_min, acc_max = acceptable_range
    if opt_min <= value <= opt_max:
        return 100
    if acc_min <= value <= acc_max:
        if value < opt_min:
            distance = opt_min - value
            max_distance = opt_min - acc_min
        else:
            distance = value - opt_max
            max_distance = acc_max - opt_max
        return max(0, 100 - (distance / max_distance * 50))
    return 0


def calculate_composite_score(row):
    total_score = 0
    total_weight = 0
    for metric, params in CRITERIA.items():
        if metric in row.index and pd.notna(row.get(metric)):
            score = score_metric(row[metric], params['optimal_range'], params['acceptable_range'])
            total_score += score * params['weight']
            total_weight += params['weight']
    if total_weight > 0:
        return total_score / total_weight
    return 0


# ============================================================================
# FILTERS
# ============================================================================

def apply_quality_filters(df):
    f = df.copy()
    f = f[f['pe_ratio'].notna() & (f['pe_ratio'] > 0)]
    f = f[f['z_score'].notna() & (f['z_score'] < -1.5)]
    f = f[f['drop_from_high_pct'] > -70]
    if 'avg_volume' in f.columns:
        f = f[f['avg_volume'].notna() & (f['avg_volume'] > 500_000)]
    if 'volatility' in f.columns:
        f = f[f['volatility'] < 150]
    return f


# ============================================================================
# COLOR CODING
# ============================================================================

# Fills
FILL_GREEN = PatternFill('solid', fgColor='C6EFCE')
FILL_YELLOW = PatternFill('solid', fgColor='FFEB9C')
FILL_RED = PatternFill('solid', fgColor='FFC7CE')
FILL_HEADER = PatternFill('solid', fgColor='2F5496')
FILL_LIGHT_GRAY = PatternFill('solid', fgColor='F2F2F2')

# Fonts
FONT_HEADER = Font(bold=True, color='FFFFFF', size=10, name='Arial')
FONT_STRONG = Font(bold=True, color='006100', size=10, name='Arial')
FONT_REVIEW = Font(color='9C5700', size=10, name='Arial')
FONT_PASS = Font(color='9C0006', size=10, name='Arial')
FONT_DEFAULT = Font(size=10, name='Arial')
FONT_WARN = Font(bold=True, color='CC0000', size=10, name='Arial')

THIN_BORDER = Border(
    bottom=Side(style='thin', color='D9D9D9')
)


def apply_formatting(ws, df):
    """Apply color coding to worksheet based on data values"""

    # Header formatting
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Build column name map
    col_map = {}
    for col_idx in range(1, ws.max_column + 1):
        col_map[ws.cell(row=1, column=col_idx).value] = col_idx

    # Row formatting
    for row_idx in range(2, ws.max_row + 1):
        # Alternating row background
        if row_idx % 2 == 0:
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = FILL_LIGHT_GRAY

        # Bottom border on all cells
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).border = THIN_BORDER
            ws.cell(row=row_idx, column=col_idx).font = FONT_DEFAULT

        # --- Tier coloring ---
        if 'tier' in col_map:
            tier_cell = ws.cell(row=row_idx, column=col_map['tier'])
            val = tier_cell.value
            if val == 'STRONG':
                tier_cell.fill = FILL_GREEN
                tier_cell.font = FONT_STRONG
            elif val == 'REVIEW':
                tier_cell.fill = FILL_YELLOW
                tier_cell.font = FONT_REVIEW
            elif val == 'PASS':
                tier_cell.fill = FILL_RED
                tier_cell.font = FONT_PASS

        # --- Opportunity score coloring ---
        if 'opportunity_score' in col_map:
            cell = ws.cell(row=row_idx, column=col_map['opportunity_score'])
            try:
                v = float(cell.value) if cell.value is not None else 0
                if v >= 70:
                    cell.fill = FILL_GREEN
                    cell.font = FONT_STRONG
                elif v >= 50:
                    cell.fill = FILL_YELLOW
                    cell.font = FONT_REVIEW
                else:
                    cell.fill = FILL_RED
                    cell.font = FONT_PASS
            except:
                pass

        # --- Earnings soon warning ---
        if 'earnings_soon' in col_map:
            cell = ws.cell(row=row_idx, column=col_map['earnings_soon'])
            if cell.value == True or str(cell.value).upper() == 'TRUE':
                cell.fill = FILL_RED
                cell.font = FONT_WARN

        # --- Earnings in window warning ---
        if 'earnings_in_window' in col_map:
            cell = ws.cell(row=row_idx, column=col_map['earnings_in_window'])
            if cell.value == True or str(cell.value).upper() == 'TRUE':
                cell.fill = FILL_RED
                cell.font = FONT_WARN

        # --- Regime coloring ---
        if 'regime' in col_map:
            cell = ws.cell(row=row_idx, column=col_map['regime'])
            val = str(cell.value) if cell.value else ''
            if val == 'Elevated':
                cell.fill = FILL_RED
                cell.font = FONT_WARN
            elif val == 'Compressed':
                cell.fill = FILL_GREEN
                cell.font = FONT_STRONG
            elif val == 'Neutral':
                cell.fill = FILL_YELLOW
                cell.font = FONT_REVIEW

        # --- Vol skew ratio coloring ---
        if 'vol_skew_ratio' in col_map:
            cell = ws.cell(row=row_idx, column=col_map['vol_skew_ratio'])
            try:
                v = float(cell.value) if cell.value is not None else 1.0
                if v >= 1.3:
                    cell.fill = FILL_RED
                    cell.font = FONT_WARN
                elif v >= 1.1:
                    cell.fill = FILL_YELLOW
                elif v < 0.9:
                    cell.fill = FILL_GREEN
            except:
                pass

        # --- Industry isolation (low count = good) ---
        if 'industry_oversold_count' in col_map:
            cell = ws.cell(row=row_idx, column=col_map['industry_oversold_count'])
            try:
                v = int(cell.value) if cell.value is not None else 0
                if v <= 1:
                    cell.fill = FILL_GREEN
                    cell.font = FONT_STRONG
                elif v >= 5:
                    cell.fill = FILL_RED
                    cell.font = FONT_WARN
            except:
                pass

        # --- Volume surge coloring ---
        if 'volume_surge' in col_map:
            cell = ws.cell(row=row_idx, column=col_map['volume_surge'])
            try:
                v = float(cell.value) if cell.value is not None else 1.0
                if v >= 1.5:
                    cell.fill = FILL_GREEN
                elif v < 0.8:
                    cell.fill = FILL_RED
            except:
                pass

    # Auto-width columns
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ''))
            for r in range(1, min(ws.max_row + 1, 50))
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 25)


# ============================================================================
# MAIN ANALYSIS
# ============================================================================

def analyze_opportunities(input_file, output_file):
    print("\n" + "="*80)
    print("OPPORTUNITY ANALYZER")
    print("="*80)

    print(f"\nLoading: {input_file}")
    df = pd.read_excel(input_file, sheet_name='All Results')
    print(f"Loaded {len(df)} stocks")

    # Quality filters
    print("\nFiltering...")
    df_filtered = apply_quality_filters(df)
    print(f"  {len(df_filtered)} passed")

    if len(df_filtered) == 0:
        print("\nNo stocks passed filters.")
        return

    # Score
    df_filtered = df_filtered.copy()
    df_filtered['opportunity_score'] = df_filtered.apply(calculate_composite_score, axis=1).round(1)
    df_filtered = df_filtered.sort_values('opportunity_score', ascending=False)

    # Tier
    df_filtered['tier'] = df_filtered['opportunity_score'].apply(
        lambda x: 'STRONG' if x >= 70 else ('REVIEW' if x >= 50 else 'PASS')
    )

    # Industry isolation flag
    if 'industry_oversold_count' in df_filtered.columns:
        df_filtered['industry_isolated'] = df_filtered['industry_oversold_count'] <= 1

    # Column order for output
    output_cols = [
        'ticker', 'opportunity_score', 'tier',
        'sector', 'industry',
        'earnings_soon', 'earnings_in_window', 'earnings_date',
        'ex_dividend_date',
        'z_score', 'distance_from_mean_pct',
        'volume_surge',
        'sector_oversold_count', 'industry_oversold_count',
        'risk_state_score', 'regime', 'vol_regime_ratio', 'tail_thickness',
        'current_price', 'recent_high', 'drop_from_high_pct',
        'p1', 'p5', 'p10', 'p25', 'p50',
        'strike_p5', 'strike_p10', 'spread_width',
        'var_95', 'cvar_95', 'var_99', 'cvar_99',
        'volatility', 'downside_vol', 'vol_skew_ratio',
        'pe_ratio', 'forward_pe',
        'avg_volume', 'signal',
    ]
    available = [c for c in output_cols if c in df_filtered.columns]
    output_df = df_filtered[available]

    # Terminal summary
    print("\n" + "="*80)
    print("TOP 20 OPPORTUNITIES")
    print("="*80)

    display_cols = ['ticker', 'opportunity_score', 'tier', 'z_score', 'pe_ratio',
                    'drop_from_high_pct', 'regime', 'vol_skew_ratio', 'earnings_soon']
    disp_avail = [c for c in display_cols if c in output_df.columns]
    print("\n" + output_df[disp_avail].head(20).to_string(index=False))

    n_strong = len(output_df[output_df['tier'] == 'STRONG'])
    n_review = len(output_df[output_df['tier'] == 'REVIEW'])
    n_pass = len(output_df[output_df['tier'] == 'PASS'])

    print(f"\n  STRONG: {n_strong}  |  REVIEW: {n_review}  |  PASS: {n_pass}")

    # Sector breakdown
    if 'sector' in df_filtered.columns:
        print("\nBy Sector:")
        sec = df_filtered.groupby('sector').agg(
            count=('ticker', 'count'),
            avg_score=('opportunity_score', 'mean')
        ).sort_values('avg_score', ascending=False)
        print(sec.head(10).to_string())

    # ========================================================================
    # WRITE EXCEL WITH FORMATTING
    # ========================================================================

    print(f"\nSaving to: {output_file}")

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        output_df.to_excel(writer, sheet_name='Ranked', index=False)

        strong = output_df[output_df['tier'] == 'STRONG']
        if len(strong) > 0:
            strong.to_excel(writer, sheet_name='Strong', index=False)

        review = output_df[output_df['tier'] == 'REVIEW']
        if len(review) > 0:
            review.to_excel(writer, sheet_name='Review', index=False)

        if 'sector' in df_filtered.columns:
            sec.to_excel(writer, sheet_name='Sector Summary')

    # Apply color coding
    wb = load_workbook(output_file)
    for sheet_name in wb.sheetnames:
        if sheet_name != 'Sector Summary':
            apply_formatting(wb[sheet_name], output_df)

    # Format sector summary header
    if 'Sector Summary' in wb.sheetnames:
        ws = wb['Sector Summary']
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = FILL_HEADER
            cell.font = FONT_HEADER

    wb.save(output_file)

    print(f"\nSheets: Ranked | Strong | Review | Sector Summary")
    print(f"Color coding applied.")
    print("="*80)


if __name__ == "__main__":
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
    else:
        input_file = INPUT_FILE

    if not Path(input_file).exists():
        print(f"\nERROR: File not found: {input_file}")
        sys.exit(1)

    analyze_opportunities(input_file, OUTPUT_FILE)